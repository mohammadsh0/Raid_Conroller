import os
import csv
import re
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from zipfile import ZipFile
from getpass import getpass
from shutil import rmtree as dirRemover
from datetime import date
from copy import copy


def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


def section_breaker(custum_list:list, start:int, end:int):
    # sourcery skip: instance-method-first-arg-name
    return(custum_list[start:end])


def num_of_lines(file_obj):
    with open (file_obj, 'r') as f_obj:
        return(len(f_obj.readlines()))


class Disk():
    def __init__(self, pd_slice):
        self.pd_parameters = {
            'devID' : '',
            'EncID' : '',
            'slotNum' : '',
            'otherError' : '',
            'mediaError' : '',
            'predictFail' : '',
        }
        self.pd_slice = pd_slice
        self.get_pd_params()

    def get_pd_params(self):
        for line in self.pd_slice:
            if 'Device Id' in line:
                self.pd_parameters['devID'] = int(line[10:].strip())
            if 'Enclosure Device ID' in line:
                self.pd_parameters['EncID'] = int(line[20:].strip())
            if 'Slot Number' in line:
                self.pd_parameters['slotNum'] = int(line[12:].strip())
            if 'Other Error Count:' in line:
                self.pd_parameters['otherError'] = line[18:].strip()
            if 'Media Error Count:' in line:
                self.pd_parameters['mediaError'] = line[18:].strip()
            if 'Predictive Failure Count:' in line:
                self.pd_parameters['predictFail'] = line[25:].strip()


class Chaos():
    def __init__(self, pdlist):
        self.pd_section_indices = []
        # List of pdlist lines
        self.pdlist_list = []
        # {disk-no: [list of pdlist lines for the Disk-#num]}
        self.pdlist_dict = {}
        self.pdlist = pdlist
        self.pdlist_last_line = num_of_lines(self.pdlist)

        # Read pdlist and extend pdlist list
        with open(self.pdlist, 'r') as f_obj:
            self.pdlist_list.extend(iter(f_obj))

    def pdlist_slicer(self):
        """Get pdlist sections line indices"""
        self.pd_section_indices = [i for i, x in enumerate(self.pdlist_list) if "Enclosure Device ID" in x]

        for i in range(len(self.pd_section_indices)):
            if i + 1 < len(self.pd_section_indices):
                self.pdlist_dict[f'Disk-{i}'] = section_breaker(self.pdlist_list, self.pd_section_indices[i], self.pd_section_indices[i+1])
            else:
                self.pdlist_dict[f'Disk-{i}'] = section_breaker(self.pdlist_list, self.pd_section_indices[i], self.pdlist_last_line)


def excel_maker(pd_params, excel_file):
    E_S = f"{pd_params['EncID']}/{pd_params['slotNum']}"
    rows = (
        # ('Device ID', 'Enc/Slot', 'Error Count', 'Value'),
        (pd_params['devID'], E_S, 'Other Error Count', pd_params['otherError']),
        (None, None, 'Media Error Count', pd_params['mediaError']),
        (None, None, 'Predictive Failure Count', pd_params['predictFail'])
    )


    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Disk_Error_Count"]
    for row in rows:
        sheet.append(row)
    sheet.append(('', '', '', ''))
    workbook.save(f"{excel_file}")


def excel_modifier():
    """Sets the column width sizes of each sheet to the appropriate value"""
    # Gets the excel file
    for file in os.scandir():
        if 'pd_temp.xlsx' in file.name:
            excel_file = file.name
            excel_file = os.path.join(os.getcwd(), excel_file)
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Disk_Error_Count"]
    center_align = Alignment(horizontal='center', vertical='center')


    dims = {}
    for row in sheet.rows:
        for cell in row:
            cell.alignment = center_align
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 4

    a_value_indexes = []
    b_value_indexes = []
    c_empty_indexes = []
    for index, column in enumerate(sheet.columns):
        if 'A' in get_column_letter(index + 1):
            a_value_indexes.extend((cell.row, cell.column) for cell in column if cell.value)
        if 'B' in get_column_letter(index + 1):
            b_value_indexes.extend((cell.row, cell.column) for cell in column if cell.value)
        if 'C' in get_column_letter(index + 1):
            c_empty_indexes.extend(cell.row - 1 for cell in column if not cell.value)

        # Formatting Columns D & E as number
        if 'D' or 'E' in get_column_letter(index + 1):
            for cell in column:
                if cell.value:
                    cell.number_format = '0'

    for indice, cell in zip(c_empty_indexes, a_value_indexes):
        sheet.merge_cells(start_row=cell[0], end_row=indice, start_column=cell[1], end_column=cell[1])
    for indice, cell in zip(c_empty_indexes, b_value_indexes):
        sheet.merge_cells(start_row=cell[0], end_row=indice, start_column=cell[1], end_column=cell[1])

    workbook.save(excel_file)


def get_rm_junk():
    msg = 'Do you want to remove files that were created by the script, after completion? (Y/N) '
    remove = input(msg)
    if remove.lower() == 'y':
        return True
    return False if remove.lower() == 'n' else (get_rm_junk())


def pdlist():
    analyze = input('Do you want to check for error counts in pdlist file? (y/n) ').lower()
    return analyze == 'y'


class LogAnalyzer():
    def __init__(self):
        self.passwd = self.get_password()
        self.rm_junk = False
        self.path = os.getcwd()
        self.rc_file = ''
        self.org_name = str(input('Enter organization name: ')).strip()
        self.id = str(input('Enter chassis ID number: '))

    def get_password(self):
        try:
            return getpass()
        except Exception as error:
            print('ERROR', error)

    def extractor(self):
        # gets RClogs###.log file path
        for file in os.scandir(self.path):
            if 'RCLogs'.lower() in file.name.lower():
                self.rc_file = file.name

        # renaming rc_file.log to rc_file.zip
        if self.rc_file[-4:].lower() != '.zip':
            os.rename(self.rc_file, f'{self.rc_file[:-4]}.zip')
            self.rc_file = f'{self.rc_file[:-4]}.zip'

        # extracting rc_file.log with the given password
        with ZipFile(
                os.path.join(os.getcwd(), self.rc_file), 'r') as zip_ref:
            zip_ref.extractall(os.getcwd(), pwd=bytes(self.passwd, 'utf-8'))

        # Unzip the extracted zip file whithin RCLogs###.zip
        for file in os.scandir():
            if '.zip' in file.name.lower() and 'RCLogs'.lower() not in file.name.lower():
                self.rc_file = file.name

        with ZipFile(os.path.join(os.getcwd(), self.rc_file), 'r') as zip_ref:
            zip_ref.extractall(os.getcwd())

        # get AllEvents and Incremental Log files and delete the rest
        file_list = []
        for file in os.scandir():
            if 'allevents' in file.name.lower():
                all_events = file.name
            elif 'incremental' in file.name.lower():
                incremental = file.name
            elif '.megalog' in file.name.lower():
                file_list.append(file.name)

        # remove Get_FwTermLog_Controller files
        for file in os.scandir():
            if 'Get_FwTermLog_Controller'.lower() in file.name.lower():
                os.remove(os.path.join(os.getcwd(), file.name))

    def convert_to_alilog(self):
        # Gets Megaraid_Incremental_Log path
        for file in os.scandir():
            if 'MegaRAID_Incremental_Log'.lower() in file.name.lower():
                inc_log_full_path = os.path.join(os.getcwd(), file.name)
                inc_log = file.name

        # Read Incremental Log File
        with open(inc_log, 'r') as f_obj:
            log_file = f_obj.readlines()

        # make a list of Inc_log lines
        # make a list of where each new message begins in the log_list
        log_list = list(log_file)
        section_indices = [i for i, x in enumerate(log_list) if "seqNum" in x]


        # Making a full line out of each message
        # Then Writing the full line to alilog file.
        complete_line = ''
        complete_lines = []
        for i in range(len(section_indices)):
            complete_line = ''
            if section_indices[i] != section_indices[-1]:
                for line in (log_list[section_indices[i]:section_indices[i+1]]):
                    complete_line += line.replace('\n', '  ')
            # Making an exception just for the last item in section_indices
            else:
                for line in (log_list[section_indices[i]:]):
                    complete_line += line.replace('\n', '  ')
            complete_line = '\n' + complete_line
            complete_lines.append(complete_line)

        # Write each line to the GetEventsToAlilog
        with open(f'GetEventsToAlilog-{inc_log}', 'w') as f_obj:
            f_obj.writelines(complete_lines)

    def oraganizer(self):
        # Get AliLog path
        path = os.getcwd()
        for file in os.scandir():
            if 'geteventstoalilog' in file.name.lower():
                ali_log = os.path.join(os.getcwd(), file.name)

        # Read Alilog lines and store them in a list
        log_lines = []
        with open(ali_log) as file_obj:
            log_lines.extend(iter(file_obj))


        # Function to look for a search term in log_lines
        # item is one of the search_dict terms and
        # logs is a list of all the RC_Log_Lines
        def searcher(item, logs):
            print(f'\tLooking for "{item.capitalize()}" in AliLog...')
            # print([line for line in logs if item.lower() in line.lower()])
            return [line for line in logs if item.lower() in line.lower()]


        # Categories to search for in log file:
        # Other Category isn't defined
        search_dict = {
            'Power state change': [],
            'Unexpected sense': [],
            'medium error': [],
            'Uncorrectable': [],
            'recovery': [],
            'Fatal firmware error': [],
            'DEGRADED': [],
            'State change on VD': [],
            'State change on PD': [],
            'Consistency Check st': [],
            'Consistency Check done': [],
            'abort': [],
            'inconsistent': [],
            'Battery': [],
            'Rebuild complete': [],
            'Rebuild failed': [],
            'Rebuild automatically started': [],
            'Rebuild started': []
            # 'ECC': []
            # 'Other': []
        }

        # Fill the value of each search_term with relevant lines from RC_Logs
        for k in search_dict:
            search_dict[k] = searcher(k, log_lines)

        # defining Other section
        other_list = []
        print('\tLooking for "Other" in AliLog...')
        for line in log_lines:
            for item in search_dict:
                if item != 'Other' and item.lower() in line.lower():
                    break
            else:
                other_list.append(line)
        search_dict['Other'] = other_list


        def alilog_list(lines):
            mylist = []
            for line in lines:
                time = re.search("Time:.([A-zZ-a]*).([A-zZ-a]*) *[0-9]* *([0-9]*:[0-9]*:[0-9]*) [0-9]*", line)
                reboot = re.search("Seconds since last reboot:.[0-9]*", line)
                description = re.search("Event Description: .* (?=Event Data)", line)
                data = re.search("Event Data: .*", line)
                if reboot:
                    mylist.append([reboot.group().strip(), description.group().strip(), data.group().strip()])
                if time:
                    mylist.append([time.group().strip(), description.group().strip(), data.group().strip()])

            return(mylist)


        for item in search_dict:
            list_of_lines = alilog_list(search_dict[item])
            with open(
                os.path.join(os.getcwd(), f'{item}.csv'), 'w', newline='') as obj:
                csv_writer = csv.writer(obj)
                csv_writer.writerows(list_of_lines)

    def excel_maker(self):
        workbook = openpyxl.Workbook()
        workbook.save(f"{self.org_name.capitalize()}-ID{self.id}-RC_Log_Analyze-{date.today().strftime('%B-%d-%Y')}.xlsx")

        # Write each csv to the excel file
        def excel_writer(sheet_name, csv_path):
            for file in os.scandir():
                if '.xlsx' in file.name:
                    self.final_file = file.name
            rows = []
            with open(csv_path) as f_obj:
                csv_reader = csv.reader(f_obj)
                rows.extend(iter(csv_reader))
            workbook = openpyxl.load_workbook(self.final_file)
            sheet = workbook.create_sheet(sheet_name)
            # column = j
            j = 1
            sheet = workbook[sheet_name]
            for i, row in enumerate(rows, start=1):
                for j, item in enumerate(row, start=1):
                    try:
                        sheet.cell(row=i, column=j, value=item)
                    except openpyxl.utils.exceptions.IllegalCharacterError:
                        sheet.cell(row=i, column=j, value=ILLEGAL_CHARACTERS_RE.sub(r'', item))
            workbook.save(f"{self.org_name.capitalize()}-ID{self.id}-RC_Log_Analyze-{date.today().strftime('%B-%d-%Y')}.xlsx")


        # Get CSV files within the folder and store them in a dict
        path_dict = {}


        def csv_files(path):
            for file in os.scandir(path):
                if ('.csv' in file.name and
                        os.path.getsize(os.path.join(path, file.name)) != 0):
                    path_dict[f'{file.name[:-4]}'] = (os.path.join(path, file.name))


        csv_files(os.getcwd())

        for file_name, file_path in path_dict.items():
            excel_writer(file_name, file_path)

        workbook = openpyxl.load_workbook(self.final_file)
        workbook.remove(workbook['Sheet'])
        workbook.save(self.final_file)

    def excel_modifier(self):
        """Sets the column width sizes of each sheet to the appropriate value"""
        # Gets the excel file
        for file in os.scandir():
            if '.xlsx' in file.name:
                excel_file = file.name
        excel_file = os.path.join(os.getcwd(), excel_file)

        #2 Copied from https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        workbook = openpyxl.load_workbook(excel_file)
        for i in workbook.sheetnames:
            sheet = workbook[i]
            sheet
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value
        workbook.save(excel_file)

    def junk_remove(self):
        csv_list = [file.name for file in os.scandir() if '.csv' in file.name.lower()]
        megalog_list = [file.name for file
                        in os.scandir() if '.megalog' in file.name.lower()]
        zip_list = [file.name for file in os.scandir() if '.zip' in file.name.lower() and 'rclogs' not in file.name.lower()]
        junk_list = csv_list + megalog_list + zip_list

        tmp_folder = os.path.join(os.getcwd(), 'tmpfiles')

        if self.rm_junk:
            for file in junk_list:
                os.remove(os.path.join(os.getcwd(), file))
        else:
            os.mkdir(os.path.join(os.getcwd(), 'tmpfiles'))
            for file in junk_list:
                os.replace(
                    os.path.join(os.getcwd(), file), os.path.join(tmp_folder, file))

    def purge(self):
        """Clears the last RC Analyze files if any are there"""
        for file in os.scandir():
            if '.xlsx' in file.name:
                os.remove(file.name)
            if 'tmpfiles' in file.name:
                print("Removing extra files from last Log analayze...")
                dirRemover(file.name)


main = LogAnalyzer()
main.rm_junk = get_rm_junk()

main.purge()

print("Extracting zip files...")
main.extractor()

print("Converting to AliLog...")
main.convert_to_alilog()

print("Organizing AliLog into CSV files...")
main.oraganizer()

print("Making the Excel output from CSV files...")
main.excel_maker()

print("Adjusting sheet column width sizes in the Excel file...")
main.excel_modifier()

if getpd := pdlist():
    print("Checking Disk Errors")
    for file in os.scandir(os.getcwd()):
        if 'pdlist' in file.name.lower():
            pdlist = file.name
    smart = Chaos(pdlist)
    smart.pdlist_slicer()

    disks = {disk: Disk(pd_data) for (disk, pd_data) in smart.pdlist_dict.items()}

    file_list = [file.name for file in os.scandir()]
    final_file = [file.name for file in os.scandir() if 'RC_Log_Analyze' in file.name]
    workbook = openpyxl.Workbook()
    workbook.create_sheet("Disk_Error_Count")
    workbook.save("pd_temp.xlsx")

    for disk in disks.values():
        excel_maker(disk.pd_parameters, "pd_temp.xlsx")
    excel_modifier()

    wb_target = openpyxl.load_workbook(final_file[0])
    target_sheet = wb_target.create_sheet('Disk_Error_Count')

    wb_source = openpyxl.load_workbook('pd_temp.xlsx')
    source_sheet = wb_source['Disk_Error_Count']

    copy_sheet(source_sheet, target_sheet)

    if 'Sheet' in wb_target.sheetnames:  # remove default sheet
        wb_target.remove(wb_target['Sheet'])

    wb_target.save(final_file[0])

print("Dealing with extra created files...")
main.junk_remove()

print("Done")
